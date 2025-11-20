import json

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


def get_models_by_layer(layer_name: str, nodes: dict, prj_name: str):
    models = {}
    for n in nodes.values():
        if n["resource_type"] == "model" and layer_name in n["path"]:
            models[n["name"]] = n["depends_on"]["nodes"][0]. \
                    replace(f"model.{prj_name}.HS_", ""). \
                    replace(f"model.{prj_name}.WRK_{n["name"].split("_")[0]}_", ""). \
                    replace(f"model.{prj_name}.", "")
    return(models)



def generate_test_plan(dbt_path: str):
    with open(f"{dbt_path}/target/manifest.json", "r") as f:
        manifest = json.load(f)
    prj_name = manifest["metadata"]["project_name"]
    nodes = manifest["nodes"]
    prj_name = manifest["metadata"]["project_name"]
    
    output_xlsx = f"{prj_name.upper()}_Test_Plan.xlsx"
    
    wb = Workbook()
    ws = wb.active
    ws.title = 'Test Plan'

    headers = [
        'TestName', 
        'Path', 
        'TestDescription', 
        'TestID', 
        'StepActionDescription', 
        'StepResultDescription', 
        'CategoriaTest', 
        'ImportanzaTest', 
        'TipoEsecuzioneTest', 
        'Business Requirements', 
        'Component', 
        'Fornitore di sviluppo'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header


    ws.cell(row=2, column=5).value = """<DIV><DIV><P>Creare le tabelle su DB BQ target. Inizializzare le tabelle BQ ad una data definita (T) usando Apollo, in modo da partire da una situazione consistente a parità di contenuto della stessa tabelle sul DB Teradata<BR/></P></DIV></DIV>"""
    ws.cell(row=2, column=6).value = """<DIV><P>Le query devono essere eseguite senza errori e le tablle devono risultare create e popolate con successo<BR/></P></DIV>"""

    ws.cell(row=3, column=5).value = """<DIV><DIV><P>Verificare che tutti i file EBCDIC riferiti al giorno (T+1) siano stati depositati sul bucket di landing<BR/></P></DIV></DIV><DIV><DIV><P>Verificare che tutti i file EBCDIC riferiti al giorno (T+1) siano stati depositati sul bucket di landing<BR/></P></DIV></DIV>"""
    ws.cell(row=3, column=6).value = """<DIV><P>I file esistono sul bucket<BR/></P></DIV>"""

    silver_models = get_models_by_layer("silver", nodes, prj_name)

    for row_num, model in enumerate(silver_models.keys(), 4):
        ws.cell(row=row_num, column=5).value = f"<DIV><DIV><P>Eseguire il SQL comparison sulla tabella {model}, del flusso {silver_models[model]}, con l'obiettivo di fare la quadratura tecnica dei dati a valle del caricamento giornaliero incrementale del giorno (T+1)<BR/></P></DIV></DIV>"
        ws.cell(row=row_num, column=6).value = f"<DIV><P>Quadratura dei dati per la tabella {model} del flusso {silver_models[model]}.<BR/></P></DIV>"

    gold_models = get_models_by_layer("gold", nodes, prj_name)
    
    print(f"silver: {len(silver_models)}, gold: {len(gold_models)}")

    for row_num, model in enumerate(gold_models.keys(), len(silver_models) + 4):
        ws.cell(row=row_num, column=5).value = f"<DIV><DIV><P>Eseguire il SQL comparison sulla vista {model}, con l'obiettivo di fare la quadratura tecnica dei dati a valle del caricamento giornaliero incrementale del giorno (T+1)<BR/></P></DIV></DIV>"
        ws.cell(row=row_num, column=6).value = f"<DIV><P>Quadratura dei dati per la vista {model}.<BR/></P></DIV>"

    for row_num in range(2, len(silver_models) + len(gold_models) + 4):
        ws.cell(row=row_num, column=1).value = f"Caricamento Flussi Hubble {prj_name}"
        ws.cell(row=row_num, column=2).value = f"/Test Plan XXXXXX - Progetto {prj_name.upper()}"
        ws.cell(row=row_num, column=7).value = "Tecnico / Architetturale"
        ws.cell(row=row_num, column=8).value = "High"
        ws.cell(row=row_num, column=9).value = "Sanity Check"
        ws.cell(row=row_num, column=12).value = "sdg"


    # Header Style: Bold, White text, Dark Blue background, Centered
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Body Style: Top-Left alignment, Text Wrap enabled
    body_alignment = Alignment(vertical="top", wrap_text=True)

    print("Applying formatting...")

    # Format Header Row
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Column Widths Configuration
    column_widths = {
        'A': 25, # TestName
        'B': 30, # Path
        'C': 35, # TestDescription
        'D': 15, # TestID
        'E': 50, # StepActionDescription
        'F': 40, # StepResultDescription
        'G': 20, # CategoriaTest
        'H': 15, # ImportanzaTest
        'I': 20, # TipoEsecuzioneTest
        'J': 20, # Business Requirements
        'K': 15, # Component
        'L': 20  # Fornitore di sviluppo
    }

    # Apply Widths and Body Alignment
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        
        # Set width if defined, else default to 20
        if col_letter in column_widths:
            ws.column_dimensions[col_letter].width = column_widths[col_letter]
        else:
            ws.column_dimensions[col_letter].width = 20
            
        # Apply text wrapping to all rows in this column (skip header)
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = body_alignment

    # 7. Save
    wb.save(output_xlsx)
    print(f"Success! Formatted file saved as: {output_xlsx}")
    wb.save(output_xlsx)
    

if __name__ == "__main__":
    generate_test_plan(dbt_path="/home/daniele/dbt/credem/hubble/xanage-edp/")
  